import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def is_cell_in_merged_range(cell, merged_cell):
    min_row, min_col, max_row, max_col = range_boundaries(merged_cell)
    return min_row <= cell.row <= max_row and min_col <= cell.column <= max_col

def read_excel_and_generate_sql(excel_file_path, sql_file_path):
    # Get the sheet
    sheet = load_workbook(excel_file_path, read_only=True).active

    # Get column names from the first row, excluding None values
    columns = [cell.value for cell in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8).__next__() if cell.value is not None]

    # Open SQL file for writing
    with open(sql_file_path, 'w') as sql_file:
        # Generate SQL INSERT statements
        for row in sheet.iter_rows(min_row=2, max_col=8, values_only=True):
            # Check for merged cells
            # merged_cells = [cell for merged_cell in sheet.merged_cells.ranges if is_cell_in_merged_range(sheet.cell(row=sheet._current_row, column=idx + 1), merged_cell)]

            values = []
            dict1={}
            for idx, cell_value in enumerate(row):
                if idx == 1:
                    cell_value = str(cell_value)
                    id = cell_value
            for idx, cell_value in enumerate(row):
                # if any(is_cell_in_merged_range(sheet.cell(row=sheet._current_row, column=idx + 1), merged_cell) for merged_cell in merged_cells):
                #     # Skip merged cells
                #     continue
                if isinstance(cell_value, str):
                    cell_value = cell_value.replace("'",'`')
                if id == 0:
                    continue
                if idx in [3, 4, 5, 6]:
                    values.append(f"{cell_value}" if cell_value is not None and isinstance(cell_value, str) else str(cell_value))
                dict1[id]=values
            for ids,answers in dict1.items():
                for i in range(len(answers)):
                    sql_statement = f"INSERT INTO answer_options (question_id, correct_answer, answer_text) VALUES ('{ids}',{'false' if i != 0 else 'true'},'{answers[i]}');"
                    print(sql_statement)
                    sql_file.write(sql_statement + '\n')
            # sql_values = ', '.join([values[i] for i in range(len(values)) if idx in [3,4,5,6] ])
            # sql_statement = f"INSERT INTO answer_options (question_id, correct_answer, answer_text) VALUES ({sql_values});"
            # sql_file.write(sql_statement + '\n')

if __name__ == "__main__":
    excel_file_path = "/Users/hothaifa/Desktop/ThinkKing/c_knowlage.xlsx"
    sql_file_path = "./ans.sql"

    read_excel_and_generate_sql(excel_file_path, sql_file_path)
    print(f"SQL file generated at: {sql_file_path}")
