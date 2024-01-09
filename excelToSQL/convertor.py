import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def is_cell_in_merged_range(cell, merged_cell):
    min_row, min_col, max_row, max_col = range_boundaries(merged_cell)
    return min_row <= cell.row <= max_row and min_col <= cell.column <= max_col

def read_excel_and_generate_sql(excel_file_path, sql_file_path):
    # Get the sheet
    sheet = load_workbook(excel_file_path, read_only=True).active

    # Get column names from the first row, excluding None values
    columns = [cell.value for cell in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8).__next__() if cell.value is not None]

    # Open SQL file for writing
    with open(sql_file_path, 'w') as sql_file:
        # Generate SQL INSERT statements
        for row in sheet.iter_rows(min_row=2, max_col=8, values_only=True):
            # Check for merged cells
            # merged_cells = [cell for merged_cell in sheet.merged_cells.ranges if is_cell_in_merged_range(sheet.cell(row=sheet._current_row, column=idx + 1), merged_cell)]

            values = []
            for idx, cell_value in enumerate(row):
                # if any(is_cell_in_merged_range(sheet.cell(row=sheet._current_row, column=idx + 1), merged_cell) for merged_cell in merged_cells):
                #     # Skip merged cells
                #     continue
                if isinstance(cell_value, str):
                    cell_value = cell_value.replace("'",'`')
                if idx==1:
                    cell_value=str(cell_value)
                values.append(f"'{cell_value}'" if cell_value is not None and isinstance(cell_value, str) else str(cell_value))

            sql_values = ', '.join([values[i] for i in range(len(values)) if i in [2,7,1] ])
            sql_statement = f"INSERT INTO questions (topic_id, c_grade_id, level,question_id,interesting_fact , question_text) VALUES (1,3,1,{sql_values});"
            sql_file.write(sql_statement + '\n')

if __name__ == "__main__":
    excel_file_path = "/Users/hothaifa/Desktop/s3/ידע עולם כתה ה 500 שאלות.xlsx"
    sql_file_path = "./sql-vav.sql"

    read_excel_and_generate_sql(excel_file_path, sql_file_path)
    print(f"SQL file generated at: {sql_file_path}")
